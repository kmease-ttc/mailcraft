"""
SDLC JIRA Metrics Exporter
============================
Runs all 13 JQL queries against the JIRA REST API and exports
results to a single Excel file — one sheet per metric.

SETUP:
1. Fill in JIRA_DOMAIN, EMAIL, API_TOKEN, and PROJECTS below
2. Run: pip install requests openpyxl
3. Run: python scripts/jira_export.py

HOW TO GET AN API TOKEN:
- Go to: https://id.atlassian.com/manage-profile/security/api-tokens
- Click "Create API token"
- Copy the token and paste it below

OUTPUT:
- Creates "SDLC_Metrics_Export.xlsx" in the same folder
- One sheet per query, all fields included
- Ready for the report builder to generate HTML email
"""

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from requests.auth import HTTPBasicAuth
from datetime import datetime
import json
import sys

# =============================================================
# CONFIGURATION — FILL THESE IN
# =============================================================

JIRA_DOMAIN = "your-company.atlassian.net"  # e.g. acme.atlassian.net
EMAIL       = "you@yourcompany.com"
API_TOKEN   = "your-api-token-here"

# Projects to query — comma-separated list used in JQL
PROJECTS    = "LSCI, LVAIRD"

OUTPUT_FILE = "SDLC_Metrics_Export.xlsx"

# =============================================================
# QUERIES — 13 queries covering the 4 pillars
# =============================================================

P = f"project IN ({PROJECTS})"

QUERIES = [
    # ── Pillar I: Delivery Performance ──
    {
        "sheet":   "1_Throughput_WorkType",
        "label":   "Throughput + Work Type Distribution",
        "metrics": "Throughput, Work Type %",
        "jql":     f'{P} AND status = Done AND resolved >= -26w ORDER BY resolved DESC',
    },
    {
        "sheet":   "2_Velocity",
        "label":   "Velocity (story points per sprint)",
        "metrics": "Velocity, Velocity Stability",
        "jql":     f'{P} AND status = Done AND resolved >= -26w AND "Story Points" > 0 ORDER BY resolved DESC',
    },
    {
        "sheet":   "3_CycleTime",
        "label":   "Cycle Time",
        "metrics": "Cycle Time",
        "jql":     f'{P} AND status = Done AND resolved >= -26w AND status WAS "In Progress" ORDER BY resolved DESC',
    },
    # ── Pillar II: Quality ──
    {
        "sheet":   "6_AllBugs",
        "label":   "All Bugs (Defect Count / Defect Density)",
        "metrics": "Defect Count, Defect Density",
        "jql":     f"{P} AND issuetype = Bug AND created >= -26w ORDER BY created DESC",
    },
    {
        "sheet":   "7_ProdBugs",
        "label":   "Production Bugs (Defect Escape Rate)",
        "metrics": "Defect Escape Rate",
        "jql":     f'{P} AND issuetype = Bug AND created >= -26w AND labels = "prod-bug" ORDER BY created DESC',
    },
    {
        "sheet":   "8_Regressions",
        "label":   "Regression Bugs",
        "metrics": "Regression Rate",
        "jql":     f'{P} AND issuetype = Bug AND labels = "regression" AND created >= -26w ORDER BY created DESC',
    },
    {
        "sheet":   "9_Rework",
        "label":   "Rework (reopened issues)",
        "metrics": "Rework Rate",
        "jql":     f"{P} AND status WAS Done AND status != Done AND updated >= -26w ORDER BY updated DESC",
    },
    # ── Pillar III: Flow Efficiency ──
    {
        "sheet":   "4_WIP",
        "label":   "Work in Progress (current snapshot)",
        "metrics": "Work in Progress",
        "jql":     f'{P} AND status = "In Progress" ORDER BY priority DESC',
    },
    {
        "sheet":   "10_UnplannedWork",
        "label":   "Unplanned Work (Bugs + Incidents completed)",
        "metrics": "Planned vs. Unplanned",
        "jql":     f"{P} AND issuetype IN (Bug, Incident) AND status = Done AND resolved >= -26w ORDER BY resolved DESC",
    },
    {
        "sheet":   "12_Blocked",
        "label":   "Blocked Issues (Dependencies)",
        "metrics": "Cross-Team Dependencies",
        "jql":     f'{P} AND status WAS "Blocked" DURING (startOfMonth(-6), now()) ORDER BY updated DESC',
    },
    # ── Pillar IV: Backlog Health ──
    {
        "sheet":   "5_BacklogReadiness",
        "label":   "Backlog Readiness",
        "metrics": "Backlog Readiness",
        "jql":     f'{P} AND status = "Ready for Development" AND sprint is EMPTY ORDER BY priority DESC',
    },
    {
        "sheet":   "11_Discovery",
        "label":   "Discovery / Investigation Work",
        "metrics": "Discovery & Investigation Ratio",
        "jql":     f'{P} AND labels IN ("spike", "research", "investigation") AND status = Done AND resolved >= -26w ORDER BY resolved DESC',
    },
    {
        "sheet":   "13_AgingBacklog",
        "label":   "Aging Backlog (no update in 30+ days)",
        "metrics": "Aging Backlog",
        "jql":     f"{P} AND status NOT IN (Done, Closed, Cancelled) AND updated <= -30d ORDER BY updated ASC",
    },
]

# Fields to pull for each issue
FIELDS = [
    "key", "summary", "issuetype", "status", "priority",
    "assignee", "reporter", "created", "updated", "resolutiondate",
    "labels", "components", "fixVersions",
    "customfield_10016",  # Story Points
    "customfield_10020",  # Sprint
    "customfield_10014",  # Epic Link
    "resolution", "timeoriginalestimate", "timespent",
]

# =============================================================
# JIRA API FUNCTIONS
# =============================================================

BASE_URL = None
AUTH     = None

def init():
    global BASE_URL, AUTH
    BASE_URL = f"https://{JIRA_DOMAIN}/rest/api/2"
    AUTH = HTTPBasicAuth(EMAIL, API_TOKEN)

def test_connection():
    """Verify credentials work before running all queries."""
    print("Testing JIRA connection...")
    resp = requests.get(f"{BASE_URL}/myself", auth=AUTH)
    if resp.status_code == 200:
        name = resp.json().get("displayName", "Unknown")
        print(f"  Connected as: {name}\n")
        return True
    else:
        print(f"  ERROR: {resp.status_code} — {resp.text}")
        print("\n  Check your JIRA_DOMAIN, EMAIL, and API_TOKEN at the top of this script.")
        return False

def run_query(jql, fields=None, max_results=1000, expand_changelog=False):
    """Run a JQL query and return all matching issues (handles pagination)."""
    if fields is None:
        fields = FIELDS

    all_issues = []
    start_at = 0
    page_size = 100

    while True:
        expand = ["names"]
        if expand_changelog:
            expand.append("changelog")
        payload = {
            "jql":        jql,
            "startAt":    start_at,
            "maxResults": page_size,
            "fields":     fields,
            "expand":     expand,
        }
        resp = requests.post(
            f"{BASE_URL}/search",
            auth=AUTH,
            headers={"Content-Type": "application/json"},
            data=json.dumps(payload),
        )

        if resp.status_code != 200:
            print(f"    Query error {resp.status_code}: {resp.text[:300]}")
            break

        data = resp.json()
        issues = data.get("issues", [])
        all_issues.extend(issues)

        total = data.get("total", 0)
        start_at += len(issues)

        if start_at >= total or len(issues) == 0:
            break
        if start_at >= max_results:
            print(f"    Reached max_results cap ({max_results}). Total available: {total}")
            break

    return all_issues

def extract_field(issue, field_name):
    """Safely extract a field value from a JIRA issue."""
    fields = issue.get("fields", {})
    val = fields.get(field_name)

    if val is None:
        return ""

    if isinstance(val, dict):
        if "displayName" in val:
            return val["displayName"]
        if "name" in val:
            return val["name"]
        if "value" in val:
            return val["value"]
        return str(val)

    if isinstance(val, list):
        parts = []
        for item in val:
            if isinstance(item, dict):
                parts.append(item.get("name") or item.get("value") or str(item))
            else:
                parts.append(str(item))
        return ", ".join(parts)

    return str(val)

def extract_sprint_name(issue):
    """Extract sprint name from the sprint custom field."""
    fields = issue.get("fields", {})
    sprint_field = fields.get("customfield_10020")
    if not sprint_field:
        return ""
    if isinstance(sprint_field, list) and sprint_field:
        sprint = sprint_field[-1]  # most recent sprint
        if isinstance(sprint, dict):
            return sprint.get("name", "")
        if isinstance(sprint, str) and "name=" in sprint:
            try:
                name_part = sprint.split("name=")[1].split(",")[0]
                return name_part
            except Exception:
                return sprint
    return str(sprint_field)

def format_date(date_str):
    """Format ISO date string to YYYY-MM-DD."""
    if not date_str:
        return ""
    try:
        dt = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return date_str

def extract_in_progress_date(issue):
    """Extract the first date the issue transitioned to 'In Progress' from changelog."""
    changelog = issue.get("changelog", {})
    histories = changelog.get("histories", [])
    # Sort oldest first
    histories.sort(key=lambda h: h.get("created", ""))
    for history in histories:
        for item in history.get("items", []):
            if item.get("field") == "status" and (item.get("toString") or "").lower() == "in progress":
                return format_date(history.get("created"))
    return ""

def compute_days_between(date1, date2):
    """Compute days between two YYYY-MM-DD date strings."""
    if not date1 or not date2:
        return ""
    try:
        d1 = datetime.strptime(date1, "%Y-%m-%d")
        d2 = datetime.strptime(date2, "%Y-%m-%d")
        return round((d2 - d1).days, 1)
    except Exception:
        return ""

# =============================================================
# EXCEL OUTPUT
# =============================================================

HEADER_COLOR = "1E2761"
HEADER_FONT  = "FFFFFF"

def style_header_row(ws):
    """Apply header styling to row 1."""
    for cell in ws[1]:
        cell.font = Font(bold=True, color=HEADER_FONT, size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="left", wrap_text=False)

def auto_width(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

def write_issues_to_sheet(ws, issues, query_label, metrics_covered):
    """Write a list of JIRA issues to a worksheet."""
    ws.append([f"Query: {query_label}"])
    ws.append([f"Metrics: {metrics_covered}"])
    ws.append([f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([f"Total results: {len(issues)}"])
    ws.append([])

    headers = [
        "Key", "Summary", "Issue Type", "Status", "Priority",
        "Assignee", "Reporter", "Sprint", "Story Points",
        "Labels", "Components", "Fix Versions",
        "Created", "Updated", "Resolved", "Started (In Progress)",
        "Cycle Time (days)", "Lead Time (days)",
        "Resolution", "Time Estimate (hrs)", "Time Spent (hrs)", "Epic"
    ]
    ws.append(headers)
    style_header_row(ws)

    for issue in issues:
        fields = issue.get("fields", {})
        estimate = fields.get("timeoriginalestimate")
        spent    = fields.get("timespent")

        created  = format_date(fields.get("created"))
        resolved = format_date(fields.get("resolutiondate"))
        started  = extract_in_progress_date(issue)

        cycle_time = compute_days_between(started, resolved) if started and resolved else ""
        lead_time  = compute_days_between(created, resolved) if created and resolved else ""

        row = [
            issue.get("key", ""),
            extract_field(issue, "summary"),
            extract_field(issue, "issuetype"),
            extract_field(issue, "status"),
            extract_field(issue, "priority"),
            extract_field(issue, "assignee"),
            extract_field(issue, "reporter"),
            extract_sprint_name(issue),
            fields.get("customfield_10016") or "",
            extract_field(issue, "labels"),
            extract_field(issue, "components"),
            extract_field(issue, "fixVersions"),
            created,
            format_date(fields.get("updated")),
            resolved,
            started,
            cycle_time,
            lead_time,
            extract_field(issue, "resolution"),
            round(estimate / 3600, 1) if estimate else "",
            round(spent / 3600, 1) if spent else "",
            fields.get("customfield_10014") or "",
        ]
        ws.append(row)

    auto_width(ws)

def write_summary_sheet(wb, results_summary):
    """Write a summary tab with query results at a glance."""
    ws = wb.create_sheet("00_SUMMARY", 0)

    ws.append(["SDLC JIRA Metrics Export"])
    ws.append([f"Projects: {PROJECTS}"])
    ws.append([f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([f"Period: Last 6 months (26 weeks)"])
    ws.append([])
    ws.append(["Sheet", "Pillar", "Metrics Covered", "Record Count", "Status"])

    for cell in ws[6]:
        cell.font = Font(bold=True, color=HEADER_FONT, size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)

    pillar_map = {
        "1_Throughput_WorkType": "I. Delivery",
        "2_Velocity":           "I. Delivery",
        "3_CycleTime":          "I. Delivery",
        "6_AllBugs":            "II. Quality",
        "7_ProdBugs":           "II. Quality",
        "8_Regressions":        "II. Quality",
        "9_Rework":             "II. Quality",
        "4_WIP":                "III. Flow",
        "10_UnplannedWork":     "III. Flow",
        "12_Blocked":           "III. Flow",
        "5_BacklogReadiness":   "IV. Backlog",
        "11_Discovery":         "IV. Backlog",
        "13_AgingBacklog":      "IV. Backlog",
    }

    for row in results_summary:
        sheet_name = row[0]
        pillar = pillar_map.get(sheet_name, "")
        ws.append([row[0], pillar, row[1], row[2], row[3]])

    auto_width(ws)

# =============================================================
# MAIN
# =============================================================

def main():
    print("=" * 55)
    print("  SDLC JIRA Metrics Exporter")
    print(f"  Projects: {PROJECTS}")
    print("=" * 55)

    if "your-company" in JIRA_DOMAIN or "your-api-token" in API_TOKEN:
        print("\nERROR: Please fill in JIRA_DOMAIN, EMAIL, and API_TOKEN")
        print("       at the top of this script before running.\n")
        sys.exit(1)

    init()

    if not test_connection():
        sys.exit(1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    results_summary = []

    for i, q in enumerate(QUERIES, 1):
        print(f"[{i:02d}/{len(QUERIES)}] {q['sheet']} ...")
        try:
            issues = run_query(q["jql"])
            count  = len(issues)
            status = "OK"
            print(f"        {count} records")
        except Exception as e:
            issues = []
            count  = 0
            status = f"ERROR: {e}"
            print(f"        FAILED — {e}")

        ws = wb.create_sheet(q["sheet"])
        write_issues_to_sheet(ws, issues, q["label"], q["metrics"])
        results_summary.append([q["sheet"], q["metrics"], count, status])

    write_summary_sheet(wb, results_summary)

    wb.save(OUTPUT_FILE)
    print(f"\nDone! Saved to: {OUTPUT_FILE}")
    print(f"Sheets: 1 summary + {len(QUERIES)} metric tabs")
    print("\nNext step: Upload this file to the Mailcraft app or Claude")
    print("           to generate the 4-pillar stakeholder report.")

if __name__ == "__main__":
    main()
